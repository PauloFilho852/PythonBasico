from openpyxl import load_workbook, Workbook
path = 'Files\\'

file_list = ['CustosAutom', 'PopulacaoPOA', 'SuperMercado']

wb_to = Workbook()

final_file = path + 'Resultado.xlsx'

print('Juntando as planilhas...')

for file_name in file_list:
    print(f'Lendo arquivo {file_name}.xlsx ...')
    wb_from = load_workbook(filename = f'{path + file_name}.xlsx')
    ws_from = wb_from[file_name]
    max_row = ws_from.max_row
    max_column = ws_from.max_column

    ws_to = wb_to.create_sheet(title = file_name)

    for i in range(1, max_row + 1):
        for j in range(1, max_column + 1):
            cell_from = ws_from.cell(row= i, column=j)
            ws_to.cell(row=i, column = j).value = cell_from.value


print('Salvando resultados...')
wb_to.remove(wb_to['Sheet'])
wb_to.save(final_file)

print('Pronto.')
