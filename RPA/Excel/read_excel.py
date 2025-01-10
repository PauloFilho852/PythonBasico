from openpyxl import load_workbook

file_name = 'Files\\read_excel.xlsx'

wb = load_workbook(filename= file_name)

ws = wb['Planilha1']

for i in range(2,5):
    nome = ws[f'A{i}'].value
    idade = ws[f'B{i}'].value
    print(f'{nome} tem {idade} anos.')
    