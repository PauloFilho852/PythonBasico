from openpyxl import load_workbook

file_name = 'Files\\read_excel.xlsx'

print('Carregando excel...')

wb = load_workbook(file_name)
ws1 = wb['Planilha1']

ws1['A7'].value = 'ALUNOS'

print('Mesclando de A7 a B7...')

ws1.merge_cells('A7:B7') # Faz a mesclagem

#ws1.unmerge_cells('A7:B7') #Retira a mesclagem

#Alterações feitas. Salvamento do arquivo.
wb.save(file_name)

print('Pronto.')    