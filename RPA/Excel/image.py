from openpyxl import Workbook
from openpyxl.drawing.image import Image

image_file_name = 'Files\\image.jpg'
xlsx_file_name = 'Files\\image.xlsx'

wb = Workbook()

ws = wb.active

ws['A1'].value = 'Você vai ver uma imagem abaixo.'

img = Image(image_file_name)

ws.add_image(img, 'A2')

ws['A15'].value = 'Olá, Python!'

wb.save(xlsx_file_name)
